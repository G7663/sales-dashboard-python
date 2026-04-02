import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import os

# --- Data load ---
df = pd.read_csv(r"C:\Users\gyans\OneDrive\Documents\WA_Fn-UseC_-Telco-Customer-Churn.csv")
df['TotalCharges'] = pd.to_numeric(df['TotalCharges'], errors='coerce')
df.dropna(inplace=True)

# --- Summary calculations ---
total_customers = len(df)
churned = df['Churn'].value_counts()['Yes']
churn_rate = round((churned / total_customers) * 100, 1)
avg_monthly = round(df['MonthlyCharges'].mean(), 2)
avg_tenure = round(df['tenure'].mean(), 1)

# --- Excel workbook ---
wb = Workbook()
ws = wb.active
ws.title = "Churn Report"

# Header
ws.merge_cells('A1:D1')
ws['A1'] = 'Customer Churn Report'
ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill('solid', fgColor='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center')

# Summary section
ws['A3'] = 'Summary'
ws['A3'].font = Font(size=12, bold=True)

headers = ['Metric', 'Value']
ws.append([])
ws.append(headers)
for cell in ws[4]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='2E75B6')

data = [
    ['Total Customers', total_customers],
    ['Churned Customers', churned],
    ['Churn Rate (%)', churn_rate],
    ['Avg Monthly Charges', avg_monthly],
    ['Avg Tenure (months)', avg_tenure],
]
for row in data:
    ws.append(row)

# Column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15

# Contract breakdown sheet
ws2 = wb.create_sheet('Contract Breakdown')
contract_data = df.groupby('Contract')['Churn'].value_counts().unstack().fillna(0)
contract_data['Churn Rate (%)'] = round(
    contract_data['Yes'] / (contract_data['Yes'] + contract_data['No']) * 100, 1)

ws2.append(['Contract Type', 'Active', 'Churned', 'Churn Rate (%)'])
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='2E75B6')

for contract, row in contract_data.iterrows():
    ws2.append([contract, int(row['No']), int(row['Yes']), row['Churn Rate (%)'] ])

for col in ['A','B','C','D']:
    ws2.column_dimensions[col].width = 20

# Save
os.makedirs('output', exist_ok=True)
wb.save(r'C:\Users\gyans\OneDrive\Documents\churn_report.xlsx')
print("Excel report saved in output folder!")